from openpyxl import Workbook
import menu
import calendar

def menu_relatorio(con):
    opcao = int(input("""
    --------------------------------------
    Escolha algumas das opções para emissão dos relátoros:
        1 - Relatório de todos os agendamentos
        2 - Relatório de agendamento por CPF
        3 - Relatório receita mensal
        5 - Sair

    Opção desejada: """))

    if (opcao == 1):
        relatorios_agedamento(con)
    elif(opcao == 2):
        relatorios_agedamento_pessoa(con)
    elif(opcao == 3):
        relatorio_receita_mensal(con)
    elif(opcao == 4):
        menu.menu(con)
    else:   
        print('Informe uma opção valida!!')
        menu.menu(con)

def relatorios_agedamento(con):
    
    book = Workbook()
    folha = book.active

    titulo = ['Nº Agendamento', 'Cpf cliente', 'Data', 'Hora', 'Status', 'Valor total']
    folha.append(titulo)

    # Procurando todos os agendamentos da clinica
    try:   
        sql = ("""Select a.cd_agendamento, c.cpf_cliente, d.data, h.hora, a.status_agendamento, a.preco_agendamento 
                        from cliente c inner join agendamento a on c.cd_cliente = a.cd_cliente 
                        inner join datas d on a.cd_agendamento = d.cd_agendamento 
                        inner join horarios h on d.cd_data = h.cd_data
                        where h.status_horario = 'O'
                        """)
        cur_relat = con.cursor()
        cur_relat.execute(sql)
        linhas = cur_relat.fetchall()
        for i in linhas:
            elementos = [i[0], i[1], i[2], i[3], i[4], i[5]]
            folha.append(elementos) 
            
        book.save('C:/Users/Lucas/Documents/GitHub/clinical-program/relatorios/relatorio-agendamento.xlsx')
        print('Relatorio emitido com sucesso!')
        cur_relat.close()
    except:
        print('Erro ao emitir o relatório de agendamentos')


def relatorios_agedamento_pessoa(con):



    # Ativando o woorkbook 
    book = Workbook()
    folha = book.active

    #Adicionando o título
    titulo = ['Nº Agendamento', 'Cpf cliente', 'Data', 'Hora', 'Status', 'Valor total']
    folha.append(titulo)

    # Achnado CPF
    cpf_cliente = str(input('Informe o cpf do cliente, que deseja emitir o relatório:  '))

    if (cpf_cliente == ''):
        print(' Por favor prencher todos os campos! ')
        menu_relatorio(con)
    if (len(cpf_cliente) == 11):
        fatia_1 = cpf_cliente[:3]
        fatia_2 = cpf_cliente[3:6]
        fatia_3 = cpf_cliente[6:9]
        fatia_4 = cpf_cliente[9:]

        cpf_formatado = "{}.{}.{}-{}".format(
        fatia_1,
        fatia_2,
        fatia_3,
        fatia_4
        ) 
    else:
        print('Informe o seu cpf completo por favor!')
        menu_relatorio(con)

    elementos = {
        'cpf': cpf_formatado,
    }
    


    # Achar se o cliente existe, fazendo uma busca pelo cpf.

    sql = ('select count(cd_cliente) from cliente where cpf_cliente = %(cpf)s')
    cursor_cpf = con.cursor()
    cursor_cpf.execute(sql, elementos)
    cpfs = cursor_cpf.fetchall()

    for i in cpfs:
        qntd_clientes = i[0]

    
    if (qntd_clientes > 1):
        print('Achamos ', qntd_clientes, ' Clientes!')
        print('Informe o codigo do cliente que deseja emitir o relatório')
        try: 
            cursor_cliente = con.cursor()
            cursor_cliente.execute('SELECT * from Cliente where cpf_cliente = %(cpf)s', elementos)
            clientes = cursor_cliente.fetchall()
            for i in clientes:
                print(i)
        except:
            print('Erro ao procurar os clientes')
    
        cd_cliente = int(input('Informe o código do cliente desejado: '))
    
    if (qntd_clientes == 1):
        print('Cpf selecionado: ', elementos['cpf'])
        opcao = str(input('Deseja mesmo emitir o relatório nesse cpf: S-SIM, N-NÃO'))
        if (opcao.upper() == 'S'):
            try: 
                cursor_cliente = con.cursor()
                cursor_cliente.execute('select * from cliente where cpf_cliente = %(cpf)s', elementos)
                cliente = cursor_cliente.fetchall()
                for i in cliente:
                    print(i)
                    cd_cliente = i[0] 
            except:
                print('Erro ao procurar o codigo do cliente')
            
            #cd_cliente = int(input('Informe o código do cliente desejado: '))
        elif (opcao.upper() == 'N'):
            print('Opcão abortada!')
            menu.menu(con)

        else:
             print('Informe uma opção valida !!')
             menu_relatorio(con)
        
    
    if (qntd_clientes == 0):
        print('Informe um cpf valido!')
        menu_relatorio(con)
    
    #Adicionando o cd do cliente encontrado na relatório
    elementos = {
        'cd': cd_cliente,
    }
    
    # Procurando todos os agendamentos da clinica
    try:   
        sql = ("""Select a.cd_agendamento, c.cpf_cliente, d.data, h.hora, a.status_agendamento, a.preco_agendamento 
                        from cliente c inner join agendamento a on c.cd_cliente = a.cd_cliente 
                        inner join datas d on a.cd_agendamento = d.cd_agendamento 
                        inner join horarios h on d.cd_data = h.cd_data
                        where h.status_horario = 'O' and c.cd_cliente = %(cd)s
                        """)
        cur_relat = con.cursor()
        cur_relat.execute(sql, elementos)
        linhas = cur_relat.fetchall()
        for i in linhas:
            elementos = [i[0], i[1], i[2], i[3], i[4], i[5]]
            folha.append(elementos) 
            
        book.save('C:/Users/Lucas/Documents/GitHub/clinical-program/relatorios/relatorio-agendamento-pessoa.xlsx')
        print('Relatorio emitido com sucesso!')
        cur_relat.close()
    except:
        print('Erro ao emitir o relatório de agendamentos')

def relatorio_receita_mensal(con):
    book = Workbook()
    folha = book.active

    titulo = ['Valor', 'Mês']
    folha.append(titulo)

    ano = int(input('Informe o ano que desejá realizar a o relatório de receita: '))
    mes = int(input('Informe o mês que desejá realizar a o relatório de receita: '))

    elementos = {
        'ano':ano,
        'mes':mes,
    }
    # Conseguindo o nome do mês a partir do numero. Obs o nome será em inglês
    nome_mes = calendar.month_name[elementos['mes']]

    # Realziando a soma da receita mensal
    sql = (""" Select sum(a.preco_agendamento)
    from agendamento a inner join datas d
    on a.cd_agendamento = d.cd_agendamento 
    where MONTH(data) = %(mes)s and YEAR(data) = %(ano)s
    """)
    cur_relat = con.cursor()
    cur_relat.execute(sql, elementos)
    linhas = cur_relat.fetchall()
    for i in linhas:
        valor = [i[0]]
        valor.append(nome_mes)

        folha.append(valor)
    book.save('C:/Users/Lucas/Documents/GitHub/clinical-program/relatorios/relatorio-receita.xlsx')
    print('Relatório emitido com sucesso!')

